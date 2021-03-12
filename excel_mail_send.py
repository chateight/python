import openpyxl
import smtplib
from email import message
import re

def mailadd_get():
    PATH = './Desktop/python'
    wb = openpyxl.load_workbook(PATH + '/py_mail_test.xlsx')
    ws = wb['シート1']
    address = []
    ptn = re.compile(r'(^[a-zA-Z0-9_.+-]+@([a-zA-Z0-9][a-zA-Z0-9-]*[a-zA-Z0-9]*\.)+[a-zA-Z]{2,}$)')
    for row in ws.iter_rows(min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        if row[0].value is not None:
            if str :=ptn.search(row[0].value ):
                address.append(str.group(0))
    return address

#
# to get address list & send using smtp server
#
to_mail = mailadd_get()
#
# メールソフトの設定と同じで良い
#
smtp_host = 'isehara-3lv.sakura.ne.jp' #サーバを指定（このサイトのサブドメイン名）
smtp_port = '587' #ポートを指定(imap指定と同じ)
smtp_account_id = 'hogehoge.sakura.ne.jp' #アカウントを指定
smtp_account_pass = 'hoge' #メールのパスワードを設定
#
# メール本体
#
from_mail = 'hogehogel@icloud.com' #送信元
msg = message.EmailMessage()
msg.set_content('本文です'); #本文を入力
msg['Subject'] = 'テキストファイルから読み出して送信' #件名を入力
msg['From'] = from_mail
msg['Bcc'] = to_mail
#
# 送信処理
#
server = smtplib.SMTP(smtp_host, smtp_port, timeout=10)
server.login(smtp_account_id, smtp_account_pass)
result = server.send_message(msg)
server.quit()