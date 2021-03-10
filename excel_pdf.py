import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
# ------------------------------------------------------------------
# 
def pdf_gen_proc(file_pdf,ws):
    doc = SimpleDocTemplate(file_pdf, pagesize=A4)
    fontname_g = "HeiseiMin-W3"
    pdfmetrics.registerFont(UnicodeCIDFont(fontname_g))
    elements = []
# data is a 2 dimentional list
    data = []
    for row in ws.rows:
        unit_aa = []    # one row data
        for col in row:
            unit_aa.append(col.value)
        data.append(unit_aa)
    tt=Table(data)
    tt.setStyle(TableStyle([('BACKGROUND',(1,1),(-2,-2),colors.white),
        ('FONT', (0, 0), (-1, -1), fontname_g, 20),
        ('GRID', (0, 0), (ws.max_column, ws.max_row), 1, colors.black),]))
    elements.append(tt)
    doc.build(elements)
# ------------------------------------------------------------------
PATH = './Desktop/python'
file_xlsx = PATH + '/py_test_mdy.xlsx'
file_pdf = PATH + '/py_test.pdf'
#
wb = load_workbook(filename = file_xlsx)
ws = wb.active
print(f'ws.max_row = {ws.max_row}')
print(f'ws.max_column = {ws.max_column}')
#
pdf_gen_proc(file_pdf,ws)
#
