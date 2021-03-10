import openpyxl

PATH = './Desktop/python'
wb = openpyxl.load_workbook(PATH + '/py_test.xlsx')
ws = wb['シート1']
key =[]
val = []
s_val = []
s_key = []
# Id and value read
for row in ws.iter_rows(min_col=3, min_row=4, max_col=ws.max_column, max_row=ws.max_row-1):
    key.append(row[0].value)
    val.append(row[1].value)
# to make sorted list
s_val = val.copy()
s_val.sort(reverse=True)
for index, value in enumerate(s_val):
    s_key.append(key[val.index(value)])
    key[val.index(value)] = []
    val[val.index(value)] = []
print(s_val, s_key)
# rewriting to excel sheet
for row in ws.iter_rows(min_col=3, min_row=4, max_col=ws.max_column, max_row=ws.max_row-1):
    row[0].value = s_key.pop()
    row[1].value = s_val.pop()
wb.save(PATH + '/py_test_mdy.xlsx')